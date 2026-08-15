"""Tests for the expired-listing bot's offline logic.

Scope note: the three network-facing modules (alerts, mls, assessor) are
not covered here, because their correctness is a question about somebody
else's live HTML and a fixture would only test my guess at it. What IS
covered is everything that decides what reaches the vendor -- name
splitting, scoring, the review/vendor split, and the append-only
workbook contract -- which is where a silent error actually costs money.

Run with:  python -m unittest discover -s tests -v
"""

import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from ct_expired_bot.models import (
    NA,
    NEEDS_MANUAL_REVIEW,
    PORTAL_UNAVAILABLE,
    AlertListing,
    Lead,
    MlsDetail,
    OwnerRecord,
    PriceChange,
    parse_money,
)
from ct_expired_bot.names import split_owner_name
from ct_expired_bot.scoring import addresses_match, count_price_drops, score_lead, sqft_mismatch_note
from ct_expired_bot.skiptrace import (
    HeaderSpecError,
    resolve_headers,
    write_review_csv,
    write_skiptrace_csv,
)
from ct_expired_bot import workbook as wb


def make_lead(**kwargs) -> Lead:
    alert = kwargs.pop("alert", AlertListing(mls_no="24000001", street_address="19 Abel Ave", town="Stamford", zip_code="06902", status="expired"))
    mls = kwargs.pop("mls", MlsDetail(mls_no=alert.mls_no))
    owner = kwargs.pop("owner", OwnerRecord())
    return Lead(alert=alert, mls=mls, owner=owner, **kwargs)


class TestNameSplitting(unittest.TestCase):
    def test_real_vgsi_names_are_last_first(self):
        """Sampled live from Vision/Stamford 2026-08-04. Reading these as
        FIRST LAST would send every one to the vendor reversed.
        """
        for printed, expected in [
            ("CARTWRIGHT ANGELA", ("ANGELA", "CARTWRIGHT")),
            ("ALI MOHAMMED", ("MOHAMMED", "ALI")),
            ("CORREA ELIZABETH", ("ELIZABETH", "CORREA")),
            ("AGURCIA REINALDO", ("REINALDO", "AGURCIA")),
            ("MELECIO JAMIE K", ("JAMIE K", "MELECIO")),
            ("COMERFORD ELAINE P", ("ELAINE P", "COMERFORD")),
            ("CAIRO MATTHEW P", ("MATTHEW P", "CAIRO")),
        ]:
            self.assertEqual(split_owner_name(printed), expected, printed)

    def test_real_vgsi_et_al_names_are_not_split(self):
        for printed in ["GOULD RICHARD ET AL", "HLEBOGIANNIS MARIA ET AL", "GUAMAN JAIME O ET AL"]:
            first, last = split_owner_name(printed)
            self.assertEqual(first, "", printed)
            self.assertEqual(last, printed)

    def test_real_vgsi_entity_names_are_not_split(self):
        for printed in ["CAROLINA K & G LLC", "390 SAN MIGUEL LLC"]:
            first, last = split_owner_name(printed)
            self.assertEqual(first, "")
            self.assertEqual(last, printed)

    def test_first_last_order_when_source_uses_it(self):
        self.assertEqual(split_owner_name("JOHN SMITH", order="first_last"), ("JOHN", "SMITH"))
        self.assertEqual(split_owner_name("JOHN A SMITH", order="first_last"), ("JOHN A", "SMITH"))

    def test_rejects_unknown_order(self):
        with self.assertRaises(ValueError):
            split_owner_name("JOHN SMITH", order="sideways")

    def test_splits_last_comma_first_regardless_of_order(self):
        self.assertEqual(split_owner_name("SMITH, JOHN"), ("JOHN", "SMITH"))
        self.assertEqual(split_owner_name("SMITH, JOHN", order="first_last"), ("JOHN", "SMITH"))

    def test_keeps_middle_initial_with_first_name(self):
        self.assertEqual(split_owner_name("SMITH, JOHN A"), ("JOHN A", "SMITH"))

    def test_suffix_rides_with_surname(self):
        self.assertEqual(split_owner_name("SMITH JOHN JR"), ("JOHN", "SMITH JR"))
        self.assertEqual(split_owner_name("JOHN SMITH JR", order="first_last"), ("JOHN", "SMITH JR"))

    def test_hyphenated_surname_still_splits(self):
        self.assertEqual(split_owner_name("SMITH-JONES MARY"), ("MARY", "SMITH-JONES"))

    def test_entities_go_whole_into_last_name(self):
        for name in [
            "SMITH FAMILY TRUST",
            "ACME PROPERTIES LLC",
            "ESTATE OF JOHN SMITH",
            "SMITH JOHN TRUSTEE",
            "123 MAIN ST REALTY INC",
        ]:
            first, last = split_owner_name(name)
            self.assertEqual(first, "", f"{name!r} must not yield a first name")
            self.assertEqual(last, name)

    def test_multiple_owners_go_whole_into_last_name(self):
        for name in ["JOHN & MARY SMITH", "SMITH JOHN AND MARY", "SMITH, JOHN, JONES, MARY"]:
            first, last = split_owner_name(name)
            self.assertEqual(first, "")
            self.assertEqual(last, name)

    def test_ambiguous_three_token_name_is_not_split(self):
        # "Is the surname BERG or VAN DER BERG?" -- refuse rather than guess.
        first, last = split_owner_name("MARY VAN DER BERG")
        self.assertEqual(first, "")
        self.assertEqual(last, "MARY VAN DER BERG")

    def test_empty_name(self):
        self.assertEqual(split_owner_name(""), ("", ""))


class TestScoring(unittest.TestCase):
    def test_counts_only_decreases(self):
        mls = MlsDetail(mls_no="1", price_history_available=True, price_history=[
            PriceChange("01/01/2026", "$500,000", "$475,000"),   # drop
            PriceChange("02/01/2026", "$475,000", "$450,000"),   # drop
            PriceChange("03/01/2026", "$450,000", "$460,000"),   # increase
        ])
        self.assertEqual(count_price_drops(mls), 2)

    def test_reduction_and_ppsf(self):
        lead = make_lead(mls=MlsDetail(
            mls_no="1", list_price_original="$500,000", list_price_final="$450,000",
            sqft_above_grade="2,000", expiration_date="07/01/2026",
        ))
        score_lead(lead, today=date(2026, 8, 4))
        self.assertEqual(lead.total_reduction_dollars, "50000")
        self.assertEqual(lead.total_reduction_pct, "10.0%")
        self.assertEqual(lead.price_per_sqft, "225.00")
        self.assertEqual(lead.days_since_expired, "34")

    def test_missing_inputs_stay_na_not_zero(self):
        lead = make_lead(mls=MlsDetail(mls_no="1", list_price_final="$450,000"))
        score_lead(lead, today=date(2026, 8, 4))
        self.assertEqual(lead.price_per_sqft, NA)
        self.assertEqual(lead.total_reduction_dollars, NA)
        self.assertEqual(lead.days_since_expired, NA)

    def test_high_from_absentee(self):
        lead = make_lead(owner=OwnerRecord(mailing_address="500 Elsewhere Rd"))
        score_lead(lead, today=date(2026, 8, 4))
        self.assertEqual(lead.absentee, "Yes")
        self.assertEqual(lead.lead_score, "High")

    def test_owner_occupied_is_not_absentee(self):
        lead = make_lead(owner=OwnerRecord(mailing_address="19 ABEL AVENUE"))
        lead.alert.street_address = "19 Abel Ave"
        score_lead(lead, today=date(2026, 8, 4))
        self.assertEqual(lead.absentee, "No")

    def test_absentee_unknown_when_mailing_missing(self):
        lead = make_lead()
        score_lead(lead, today=date(2026, 8, 4))
        self.assertEqual(lead.absentee, NA)
        self.assertNotEqual(lead.lead_score, "High")

    def test_high_from_two_price_drops(self):
        lead = make_lead(mls=MlsDetail(mls_no="1", price_history_available=True, price_history=[
            PriceChange("01/01/2026", "$500,000", "$475,000"),
            PriceChange("02/01/2026", "$475,000", "$450,000"),
        ]))
        score_lead(lead, today=date(2026, 8, 4))
        self.assertEqual(lead.lead_score, "High")

    def test_high_from_dom(self):
        lead = make_lead(mls=MlsDetail(mls_no="1", cumulative_dom="180"))
        score_lead(lead, today=date(2026, 8, 4))
        self.assertEqual(lead.lead_score, "High")

    def test_medium_and_low_boundaries(self):
        medium = make_lead(mls=MlsDetail(mls_no="1", cumulative_dom="60"))
        score_lead(medium, today=date(2026, 8, 4))
        self.assertEqual(medium.lead_score, "Medium")

        upper = make_lead(mls=MlsDetail(mls_no="1", cumulative_dom="119"))
        score_lead(upper, today=date(2026, 8, 4))
        self.assertEqual(upper.lead_score, "Medium")

        low = make_lead(mls=MlsDetail(mls_no="1", cumulative_dom="59"))
        score_lead(low, today=date(2026, 8, 4))
        self.assertEqual(low.lead_score, "Low")

    def test_address_abbreviation_folding(self):
        self.assertTrue(addresses_match("123 Main Street", "123 MAIN ST"))
        self.assertTrue(addresses_match("45 North Rd", "45 N ROAD"))
        self.assertFalse(addresses_match("123 Main St", "124 Main St"))
        self.assertIsNone(addresses_match("123 Main St", NA))

    def test_sqft_mismatch_note(self):
        mls = MlsDetail(mls_no="1", sqft_above_grade="2,000")
        self.assertIsNone(sqft_mismatch_note(mls, OwnerRecord(assessor_living_area="1,900")))
        self.assertIsNotNone(sqft_mismatch_note(mls, OwnerRecord(assessor_living_area="900")))

    def test_parse_money_never_returns_zero_for_missing(self):
        self.assertIsNone(parse_money(NA))
        self.assertIsNone(parse_money(""))
        self.assertIsNone(parse_money("n/a"))
        self.assertEqual(parse_money("$1,234.50"), 1234.50)


class TestSkiptraceCsv(unittest.TestCase):
    def setUp(self):
        self.dir = Path(tempfile.mkdtemp())

    def test_review_rows_never_reach_vendor_file(self):
        clean = make_lead(owner=OwnerRecord(owner_name="SMITH JOHN", mailing_address="19 Abel Ave"))
        flagged = make_lead(
            alert=AlertListing(mls_no="24000002", street_address="5 Oak St", town="Stamford"),
            owner=OwnerRecord(status=NEEDS_MANUAL_REVIEW, owner_name=NEEDS_MANUAL_REVIEW, notes=["two parcels matched"]),
        )
        unavailable = make_lead(
            alert=AlertListing(mls_no="24000003", street_address="7 Elm St", town="Hartford"),
            owner=OwnerRecord(status=PORTAL_UNAVAILABLE, owner_name=PORTAL_UNAVAILABLE),
        )
        leads = [clean, flagged, unavailable]

        vendor = self.dir / "skiptrace.csv"
        review = self.dir / "review.csv"
        self.assertEqual(write_skiptrace_csv(vendor, leads), 1)
        self.assertEqual(write_review_csv(review, leads), 2)

        with vendor.open() as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Owner First Name"], "JOHN")
        self.assertEqual(rows[0]["Owner Last Name"], "SMITH")
        self.assertEqual(rows[0]["Property State"], "CT")
        blob = vendor.read_text()
        self.assertNotIn(NEEDS_MANUAL_REVIEW, blob)
        self.assertNotIn(PORTAL_UNAVAILABLE, blob)

    def test_na_never_written_to_vendor_file(self):
        lead = make_lead(owner=OwnerRecord(owner_name="SMITH JOHN"))  # mailing fields all NA
        vendor = self.dir / "skiptrace.csv"
        write_skiptrace_csv(vendor, [lead])
        with vendor.open() as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(rows[0]["Mailing Address"], "")

    def test_custom_headers_positional(self):
        headers = resolve_headers(csv_headers="F,L,Addr,City,St,Zip,MAddr,MCity,MSt,MZip")
        vendor = self.dir / "custom.csv"
        write_skiptrace_csv(vendor, [make_lead(owner=OwnerRecord(owner_name="SMITH JOHN"))], headers=headers)
        first_line = vendor.read_text().splitlines()[0]
        self.assertEqual(first_line, "F,L,Addr,City,St,Zip,MAddr,MCity,MSt,MZip")

    def test_wrong_header_count_refuses(self):
        with self.assertRaises(HeaderSpecError):
            resolve_headers(csv_headers="First,Last,Address")


class TestWorkbook(unittest.TestCase):
    def setUp(self):
        self.path = Path(tempfile.mkdtemp()) / "CT-Expired-Master.xlsx"

    def test_creates_three_tabs_with_manual_columns(self):
        self.assertTrue(wb.ensure_workbook(self.path))
        from openpyxl import load_workbook

        book = load_workbook(self.path)
        self.assertEqual(set(book.sheetnames), {"Leads", "Town Portals", "Meta"})
        headers = [c.value for c in book["Leads"][1]]
        for manual in wb.MANUAL_COLUMNS:
            self.assertIn(manual, headers)

    def test_append_preserves_manual_edits(self):
        wb.ensure_workbook(self.path)
        wb.append_leads(self.path, [make_lead(owner=OwnerRecord(owner_name="SMITH JOHN"))])

        from openpyxl import load_workbook

        book = load_workbook(self.path)
        sheet = book["Leads"]
        headers = {c.value: c.column for c in sheet[1]}
        sheet.cell(row=2, column=headers["Phone 1"], value="203-555-0100")
        sheet.cell(row=2, column=headers["My Notes"], value="left voicemail")
        book.save(self.path)

        second = make_lead(alert=AlertListing(mls_no="24000099", street_address="8 Pine St", town="Stamford"))
        wb.append_leads(self.path, [second])

        book = load_workbook(self.path)
        sheet = book["Leads"]
        self.assertEqual(sheet.cell(row=2, column=headers["Phone 1"]).value, "203-555-0100")
        self.assertEqual(sheet.cell(row=2, column=headers["My Notes"]).value, "left voicemail")
        self.assertEqual(sheet.cell(row=3, column=headers["MLS #"]).value, "24000099")
        self.assertIsNone(sheet.cell(row=3, column=headers["Phone 1"]).value)

    def test_existing_mls_numbers_drives_dedupe(self):
        wb.ensure_workbook(self.path)
        self.assertEqual(wb.existing_mls_numbers(self.path), set())
        wb.append_leads(self.path, [make_lead()])
        self.assertEqual(wb.existing_mls_numbers(self.path), {"24000001"})

    def test_append_survives_reordered_columns(self):
        wb.ensure_workbook(self.path)
        from openpyxl import load_workbook

        book = load_workbook(self.path)
        book["Leads"].insert_cols(1)
        book["Leads"].cell(row=1, column=1, value="My Own Column")
        book.save(self.path)

        wb.append_leads(self.path, [make_lead(owner=OwnerRecord(owner_name="SMITH JOHN"))])
        book = load_workbook(self.path)
        sheet = book["Leads"]
        headers = {c.value: c.column for c in sheet[1]}
        self.assertEqual(sheet.cell(row=2, column=headers["MLS #"]).value, "24000001")
        self.assertEqual(sheet.cell(row=2, column=headers["Owner Name"]).value, "SMITH JOHN")

    def test_last_run_roundtrip(self):
        wb.ensure_workbook(self.path)
        self.assertEqual(wb.read_last_run(self.path), "")
        wb.write_last_run(self.path, "2026-08-04T12:00:00+00:00")
        self.assertEqual(wb.read_last_run(self.path), "2026-08-04T12:00:00+00:00")

    def test_portals_roundtrip(self):
        from ct_expired_bot.models import TownPortal

        wb.ensure_workbook(self.path)
        wb.upsert_portals(self.path, [TownPortal("Stamford", "https://gis.vgsi.com/stamfordct/", "Vision", "ok")])
        registry = wb.read_portals(self.path)
        self.assertEqual(registry["Stamford"].platform, "Vision")
        # Upsert must update in place, not duplicate the row.
        wb.upsert_portals(self.path, [TownPortal("Stamford", "https://example.com", "Custom", "changed")])
        registry = wb.read_portals(self.path)
        self.assertEqual(len(registry), 1)
        self.assertEqual(registry["Stamford"].platform, "Custom")


if __name__ == "__main__":
    unittest.main()


class TestGraphMail(unittest.TestCase):
    """Offline coverage of the Graph payload handling.

    The auth flow and the live query are not tested here -- they need a
    real tenant. What is tested is the part that turns Microsoft's JSON
    into alert rows, and the $filter string, since a malformed filter
    silently returns the whole mailbox instead of erroring.
    """

    def test_message_to_listings_parses_body(self):
        from ct_expired_bot.graph_mail import message_to_listings

        message = {
            "id": "AAMkAGI2abc",
            "subject": "SmartMLS Listing Alert",
            "receivedDateTime": "2026-08-03T11:30:00Z",
            "from": {"emailAddress": {"address": "alerts@smartmls.com"}},
            "body": {"contentType": "html", "content": (
                "<table><tr><td>Expired</td><td>24012345</td>"
                "<td>19 Abel Avenue</td><td>06906</td></tr></table>"
            )},
        }
        listings = message_to_listings(message)
        self.assertEqual(len(listings), 1)
        self.assertEqual(listings[0].mls_no, "24012345")
        self.assertEqual(listings[0].status, "expired")
        self.assertEqual(listings[0].zip_code, "06906")
        self.assertEqual(listings[0].received_at, "2026-08-03T11:30:00Z")
        self.assertIn("alerts@smartmls.com", listings[0].source_email)

    def test_subject_filter_is_applied_client_side(self):
        from ct_expired_bot.graph_mail import message_to_listings

        message = {
            "id": "x", "subject": "Lunch tomorrow?",
            "receivedDateTime": "2026-08-03T11:30:00Z",
            "body": {"content": "<table><tr><td>Expired</td><td>24012345</td>"
                                "<td>19 Abel Avenue</td></tr></table>"},
        }
        self.assertEqual(message_to_listings(message, subject_filter="SmartMLS"), [])
        self.assertEqual(len(message_to_listings(message, subject_filter="")), 1)

    def test_empty_body_yields_nothing(self):
        from ct_expired_bot.graph_mail import message_to_listings

        self.assertEqual(message_to_listings({"id": "x", "subject": "SmartMLS"}), [])

    def test_filter_string_shape(self):
        from datetime import datetime, timezone
        from ct_expired_bot.graph_mail import _messages_url

        url = _messages_url(datetime(2026, 8, 1, tzinfo=timezone.utc), "alerts@smartmls.com")
        self.assertIn("receivedDateTime+ge+2026-08-01T00%3A00%3A00Z", url)
        self.assertIn("from%2FemailAddress%2Faddress+eq+%27alerts%40smartmls.com%27", url)
        # Subject must NOT be in $filter -- Graph rejects contains() there.
        # ($filter arrives percent-encoded as %24filter.)
        filter_clause = url.split("%24filter=")[-1]
        self.assertNotIn("subject", filter_clause)
        self.assertIn("receivedDateTime", filter_clause)

    def test_filter_omitted_when_no_constraints(self):
        from ct_expired_bot.graph_mail import _messages_url

        self.assertNotIn("%24filter", _messages_url(None, None))

    def test_single_quote_in_sender_is_escaped(self):
        from ct_expired_bot.graph_mail import _messages_url

        url = _messages_url(None, "o'brien@example.com")
        self.assertIn("o%27%27brien", url)


# Reconstructed from a screenshot of a real "Updated Matches for Matt
# Expired" alert (2026-08-13). The TEXT and field order are what was
# observed; the exact tags are not, so the parser is anchored on the
# "MLS#:" label and on line structure rather than on markup.
REAL_ALERT_HTML = """
<div><p>Hi Matt,</p>
<p>I have found 1 property that matches your search criteria.</p>
<p>This listing is either new to the market or an existing listing that
now matches your search criteria.</p>
<p>Sincerely,<br/>Matthew Ramos</p>
<h3>See below for a summary</h3>
<table><tr><td>
  <span>EXPIRED</span> <span>Expired</span>
  <div>$320,000</div>
  <a href="https://smartmls-portal.connectmls.com/x">701 Bucks Hill Road</a>
  <div>Waterbury, CT 06704</div>
  <div>MLS#: 24139663</div>
  <div>3 Beds | 2 Baths | 1,116 SqFt | Built 1993</div>
  <a href="#">View Details</a>
</td></tr></table>
<a href="#">View All Listings</a></div>
"""


class TestRealAlertLayout(unittest.TestCase):
    def test_parses_the_observed_alert(self):
        from ct_expired_bot.alerts import parse_alert_html

        listings = parse_alert_html(REAL_ALERT_HTML, source="test", received_at="2026-08-13T04:16:00Z")
        self.assertEqual(len(listings), 1)
        listing = listings[0]
        self.assertEqual(listing.mls_no, "24139663")
        self.assertEqual(listing.street_address, "701 Bucks Hill Road")
        self.assertEqual(listing.town, "Waterbury")
        self.assertEqual(listing.zip_code, "06704")
        self.assertEqual(listing.status, "expired")

    def test_sqft_is_not_mistaken_for_an_mls_number(self):
        """'1,116 SqFt' -> 1116 would match a bare \\d{4,} pattern."""
        from ct_expired_bot.alerts import parse_alert_html

        listings = parse_alert_html(REAL_ALERT_HTML)
        self.assertEqual([l.mls_no for l in listings], ["24139663"])

    def test_price_line_is_not_mistaken_for_the_address(self):
        from ct_expired_bot.alerts import parse_alert_html

        listing = parse_alert_html(REAL_ALERT_HTML)[0]
        self.assertNotIn("320", listing.street_address)
        self.assertNotIn("$", listing.street_address)

    def test_multiple_listings_in_one_alert(self):
        from ct_expired_bot.alerts import parse_alert_html

        second = REAL_ALERT_HTML.replace("</table>", """
        <tr><td><span>Withdrawn</span><div>$450,000</div>
        <a>12 Maple Street</a><div>Naugatuck, CT 06770</div>
        <div>MLS#: 24139999</div>
        <div>4 Beds | 3 Baths | 2,200 SqFt | Built 1960</div></td></tr></table>""")
        listings = parse_alert_html(second)
        self.assertEqual(len(listings), 2)
        self.assertEqual([l.mls_no for l in listings], ["24139663", "24139999"])
        self.assertEqual(listings[1].town, "Naugatuck")
        self.assertEqual(listings[1].street_address, "12 Maple Street")
        self.assertEqual(listings[1].status, "withdrawn")

    def test_active_listing_is_ignored(self):
        from ct_expired_bot.alerts import parse_alert_html

        active = REAL_ALERT_HTML.replace("<span>EXPIRED</span> <span>Expired</span>", "<span>Active</span>")
        self.assertEqual(parse_alert_html(active), [])

    def test_dedupe_across_repeated_alerts(self):
        from ct_expired_bot.alerts import dedupe, parse_alert_html

        twice = parse_alert_html(REAL_ALERT_HTML) + parse_alert_html(REAL_ALERT_HTML)
        self.assertEqual(len(dedupe(twice)), 1)


class TestGraphFolderScoping(unittest.TestCase):
    def test_folder_scopes_the_message_query(self):
        from ct_expired_bot.graph_mail import _messages_url

        scoped = _messages_url(None, None, "AAMkFOLDER")
        self.assertIn("/me/mailFolders/AAMkFOLDER/messages", scoped)
        self.assertNotIn("/me/messages", _messages_url(None, None, "AAMkFOLDER"))

    def test_no_folder_reads_whole_mailbox(self):
        from ct_expired_bot.graph_mail import _messages_url

        self.assertIn("/me/messages", _messages_url(None, None, None))

    def test_default_subject_filter_does_not_drop_real_alerts(self):
        """The observed subject is 'Updated Matches for Matt Expired -
        400K less' and the sender is the agent, not SmartMLS. A default
        subject filter of 'SmartMLS' matched zero real messages.
        """
        from ct_expired_bot.graph_mail import message_to_listings

        message = {
            "id": "x",
            "subject": "Updated Matches for Matt Expired - 400K less",
            "receivedDateTime": "2026-08-13T04:16:00Z",
            "from": {"emailAddress": {"address": "matt.ramos@newerainvesting.com"}},
            "body": {"content": REAL_ALERT_HTML},
        }
        self.assertEqual(len(message_to_listings(message)), 1)
        self.assertEqual(message_to_listings(message, subject_filter="SmartMLS"), [])
        self.assertEqual(len(message_to_listings(message, subject_filter="Updated Matches")), 1)


class TestListingUrlTemplate(unittest.TestCase):
    def test_template_substitutes_mls_number(self):
        from ct_expired_bot.mls import listing_url

        self.assertEqual(
            listing_url("24139663", "https://host/listing?mls={mls_no}"),
            "https://host/listing?mls=24139663",
        )

    def test_missing_template_raises_rather_than_guessing(self):
        from ct_expired_bot.mls import ListingUrlUnknown, listing_url

        with self.assertRaises(ListingUrlUnknown):
            listing_url("24139663", "")

    def test_template_without_placeholder_raises(self):
        from ct_expired_bot.mls import ListingUrlUnknown, listing_url

        with self.assertRaises(ListingUrlUnknown):
            listing_url("24139663", "https://host/listing")


class TestConnectMlsApi(unittest.TestCase):
    """Pinned to a real /api/shared-link response captured 2026-08-13.

    The fixture is the genuine payload for MLS 24119274 (105 Camptown
    Ave, Derby) with only the photo arrays blanked out.
    """

    @classmethod
    def setUpClass(cls):
        import json as _json

        path = Path(__file__).parent / "fixtures_shared_link.json"
        cls.payload = _json.loads(path.read_text())

    def test_maps_the_real_payload(self):
        from ct_expired_bot.connectmls_api import listing_from_payload

        detail = listing_from_payload(self.payload["Listings"][0])
        self.assertEqual(detail.mls_no, "24119274")
        self.assertEqual(detail.list_price_final, "480000")
        self.assertEqual(detail.list_price_original, "520000")
        self.assertEqual(detail.beds, "4")
        self.assertEqual(detail.full_baths, "2")
        self.assertEqual(detail.half_baths, "0")
        self.assertEqual(detail.sqft_above_grade, "1496")
        self.assertEqual(detail.year_built, "1884")
        self.assertEqual(detail.town, "Derby")
        self.assertEqual(detail.zip_code, "06418")
        self.assertIn("Multi-Family", detail.property_type)
        self.assertIn("SELLER CREDIT", detail.public_remarks)

    def test_fields_the_api_omits_stay_na(self):
        from ct_expired_bot.connectmls_api import listing_from_payload

        detail = listing_from_payload(self.payload["Listings"][0])
        for missing in (
            detail.days_on_market, detail.cumulative_dom,
            detail.list_date, detail.expiration_date,
            detail.listing_agent, detail.brokerage,
        ):
            self.assertEqual(missing, NA)
        self.assertFalse(detail.price_history_available)

    def test_remarks_truncated_to_200(self):
        from ct_expired_bot.connectmls_api import listing_from_payload

        detail = listing_from_payload(self.payload["Listings"][0])
        self.assertLessEqual(len(detail.public_remarks), 200)

    def test_status_read_from_payload(self):
        from ct_expired_bot.connectmls_api import listing_status

        self.assertEqual(listing_status(self.payload["Listings"][0]), "Expired")

    def test_uuid_extracted_from_shared_link_url(self):
        from ct_expired_bot.connectmls_api import shared_link_uuid

        url = ("https://smartmls-portal.connectmls.com/shared-link/"
               "derby-multi-family-for-sale-for-sale-105-camptown-ave/"
               "c8df80ed-23a0-4a45-87ff-03e58ec37b1b")
        self.assertEqual(shared_link_uuid(url), "c8df80ed-23a0-4a45-87ff-03e58ec37b1b")
        self.assertIsNone(shared_link_uuid("https://example.com/nope"))


class TestUnknownVsZeroPriceDrops(unittest.TestCase):
    """A 520k -> 480k listing with no SalesHistory must not read as
    'never dropped the price'. That is the shape every shared-link
    listing arrives in.
    """

    def _derby_lead(self):
        return make_lead(mls=MlsDetail(
            mls_no="24119274",
            list_price_original="520000", list_price_final="480000",
            sqft_above_grade="1496",
        ))

    def test_drops_unknown_not_zero(self):
        lead = self._derby_lead()
        score_lead(lead, today=date(2026, 8, 14))
        self.assertIsNone(lead.price_drops)

    def test_reduction_still_computed_exactly(self):
        lead = self._derby_lead()
        score_lead(lead, today=date(2026, 8, 14))
        self.assertEqual(lead.total_reduction_dollars, "40000")
        self.assertEqual(lead.total_reduction_pct, "7.7%")

    def test_unknown_drops_with_reduction_scores_medium(self):
        lead = self._derby_lead()
        score_lead(lead, today=date(2026, 8, 14))
        self.assertEqual(lead.lead_score, "Medium")

    def test_unknown_drops_without_reduction_scores_low(self):
        lead = make_lead(mls=MlsDetail(
            mls_no="1", list_price_original="500000", list_price_final="500000",
        ))
        score_lead(lead, today=date(2026, 8, 14))
        self.assertEqual(lead.lead_score, "Low")

    def test_known_zero_drops_is_still_zero(self):
        lead = make_lead(mls=MlsDetail(mls_no="1", price_history_available=True))
        score_lead(lead, today=date(2026, 8, 14))
        self.assertEqual(lead.price_drops, 0)

    def test_workbook_renders_unknown_drops_as_na(self):
        from ct_expired_bot.workbook import BOT_COLUMNS

        getter = dict(BOT_COLUMNS)["Price Drops"]
        lead = self._derby_lead()
        score_lead(lead, today=date(2026, 8, 14))
        self.assertEqual(getter(lead), NA)


# Transcribed from a real connectMLS "Listing" tab, 2026-08-14:
# 302 Dover Street, Bridgeport CT 06610, MLS 24171033 area, $324,900.
# Prices deliberately made DIFFERENT from the real row (which had all
# three equal) so the substring-collision bug would actually show.
REAL_LISTING_TAB = """Showing Instructions: Use Showing time and go
Lock Box Description: Combo
Lock Box Location: Front Door
Directions: Boston Ave to dover ST
Sign: No
Owner Name / Phone: Whitheld /
Occupied By: Vacant
Bank Owned: No
Listing Contract Type: Exclusive Right to Sell Listing Agreement
Service Type: Full Service
Potential Short Sale / Comments: No /
Acceptable Financing: FHA, VA, CHFA
Date Available: IMmediatly
Listing Agent/Broker Information
List Agent: Arbis Faustin (4119)
List Office: Faustin Realty Group (4104)
Marketing History
List Price: $324,900
Previous List Price: $349,900
Original List Price: $379,900
Price Last Updated: 06/15/2026
List Price as % of Assessed Value: 405%
Entered in MLS: 04/28/2026
Start Marketing Date: 04/27/2026
Listing Last Updated: 07/28/2026
Expiration Date: 07/27/2026
"""


class TestRealListingTabLabels(unittest.TestCase):
    """Pinned to labels read off a live connectMLS Listing tab."""

    def _fields(self):
        from ct_expired_bot.mls import LABELS, extract_fields

        return extract_fields(REAL_LISTING_TAB, LABELS)

    def test_list_price_is_not_the_original_list_price(self):
        """'List Price' is a substring of 'Original List Price' and
        'Previous List Price'. An unanchored search returns the wrong one.
        """
        fields = self._fields()
        self.assertEqual(fields["list_price_final"], "$324,900")
        self.assertEqual(fields["list_price_original"], "$379,900")
        self.assertEqual(fields["previous_list_price"], "$349,900")

    def test_marketing_history_dates(self):
        fields = self._fields()
        self.assertEqual(fields["expiration_date"], "07/27/2026")
        self.assertEqual(fields["list_date"], "04/28/2026")
        self.assertEqual(fields["price_last_updated"], "06/15/2026")

    def test_agent_and_office(self):
        fields = self._fields()
        self.assertEqual(fields["listing_agent"], "Arbis Faustin (4119)")
        self.assertEqual(fields["brokerage"], "Faustin Realty Group (4104)")

    def test_blank_value_does_not_swallow_next_field(self):
        """connectMLS leaves 'Price Last Updated:' empty when a listing
        never changed price -- it must not absorb the following line.
        """
        from ct_expired_bot.mls import LABELS, extract_fields

        blank = REAL_LISTING_TAB.replace("Price Last Updated: 06/15/2026", "Price Last Updated:")
        fields = extract_fields(blank, LABELS)
        self.assertNotIn("price_last_updated", fields)
        self.assertEqual(fields["expiration_date"], "07/27/2026")

    def test_expd_status_code_recognised(self):
        """The results grid prints 'EXPD', not 'Expired'."""
        from ct_expired_bot.models import EXPIRED_STATUSES

        self.assertIn("expd", EXPIRED_STATUSES)
        self.assertIn("expired", EXPIRED_STATUSES)
