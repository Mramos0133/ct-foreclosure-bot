"""Builds the ranked, multi-sheet Excel workbook from stored CaseResults.

Kept as a separate export step rather than folded into the incremental
per-case CSV writer, because ranking/sorting needs every matched case's
data at once (you can't sort a sheet as you go, only after everything is
in) -- the incremental writer stays for a resilient running record, this
reads the checkpoint's full case_results table and produces the final,
sorted, bucketed deliverable. Safe to re-run at any point during or after
a crawl to get an up-to-date snapshot.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .models import CaseResult

COLUMNS = [
    ("Town", lambda r: r.town),
    ("Street Address", lambda r: r.street_address),
    ("Zip Code", lambda r: r.zip_code),
    ("Docket No.", lambda r: r.docket_no),
    ("Case Caption", lambda r: r.case_caption),
    ("Summary of Case", lambda r: r.case_summary),
    ("Focus Contact", lambda r: r.focus_contact),
    ("Owner Name", lambda r: r.owner_name),
    ("Owner Phone Number", lambda r: r.owner_phone),
    ("Owner Address", lambda r: r.owner_address),
    ("Owner Address Source Doc(s)", lambda r: r.owner_address_source_urls),
    ("New Decision Maker", lambda r: r.new_decision_maker_name),
    ("New Decision Maker Phone Number", lambda r: r.new_decision_maker_phone),
    ("New Decision Maker Source Doc", lambda r: r.new_decision_maker_source_url),
    ("Motion Type Found", lambda r: "; ".join(r.motion_types_found)),
    ("Judgment Granted (Y/N)", lambda r: "Y" if r.judgment_granted else "N"),
    ("Non-Appearing (Y/N)", lambda r: "Y" if r.non_appearing else "N"),
    ("Key Date", lambda r: r.key_date or ""),
    ("Key Date Type", lambda r: r.key_date_label or ""),
    ("Days to Key Date", lambda r: r.days_to_key_date if r.days_to_key_date is not None else ""),
    ("On Auction Site (Y/N)", lambda r: "Y" if r.on_auction_site else "N"),
    ("Bankruptcy Chapter", lambda r: r.bankruptcy_chapter or ""),
    ("Continuance Count", lambda r: r.continuance_count),
    ("Warm-Cold Sub-Flag (Y/N)", lambda r: "Y" if r.warm_cold_subflag else "N"),
    ("Total Debt", lambda r: f"{r.total_debt:.2f}" if r.total_debt is not None else ""),
    ("Appraised Value", lambda r: f"{r.appraised_value:.2f}" if r.appraised_value is not None else ""),
    ("Encumbrances Subsequent to Plaintiff's Lien", lambda r: r.encumbrances_subsequent_itemized),
    ("Debt + Encumbrances / Appraised Value", lambda r: _short_sale_ratio_display(r)),
    ("Attorney Fees", lambda r: r.attorney_fees or ""),
    ("Motion for Default - Failure to Appear (Y/N)", lambda r: "Y" if r.default_failure_to_appear else "N"),
    ("Bankruptcy Stay + Reopened (Y/N)", lambda r: "Y" if r.bankruptcy_stay_reopened else "N"),
    ("Bankruptcy Supporting Text", lambda r: r.bankruptcy_supporting_text),
    ("Case Detail URL", lambda r: r.case_detail_url),
    ("Foreclosure Worksheet Doc URL", lambda r: r.worksheet_doc_url or ""),
]

SHEET_ORDER = ["HOT", "WARM", "COLD", "POTENTIAL_SHORT_SALE", "UNCLASSIFIED"]


def _short_sale_ratio(r: CaseResult) -> float | None:
    if r.total_debt is None or r.appraised_value is None or r.appraised_value <= 0:
        return None
    return (r.total_debt + (r.encumbrances_subsequent_to_lien or 0)) / r.appraised_value


def _short_sale_ratio_display(r: CaseResult) -> str:
    ratio = _short_sale_ratio(r)
    return f"{ratio:.1%}" if ratio is not None else ""


def _sort_key_hot(r: CaseResult):
    # Soonest key date first; cases with no date found sort last.
    return (r.days_to_key_date is None, r.days_to_key_date if r.days_to_key_date is not None else 0)


def _sort_key_warm(r: CaseResult):
    return (r.days_to_key_date is None, r.days_to_key_date if r.days_to_key_date is not None else 0)


def _sort_key_cold(r: CaseResult):
    # The "worth a second look" sub-segment first, then by how many times
    # extended (more extensions = has been sitting/shopped longer).
    return (not r.warm_cold_subflag, -r.continuance_count)


def _sort_key_unclassified(r: CaseResult):
    return (r.town, r.docket_no)


def _sort_key_short_sale(r: CaseResult):
    # Highest debt-to-value ratio first -- least equity cushion, most
    # urgent to price as a short sale.
    ratio = _short_sale_ratio(r)
    return (ratio is None, -(ratio or 0))


SORT_KEYS = {
    "HOT": _sort_key_hot,
    "WARM": _sort_key_warm,
    "COLD": _sort_key_cold,
    "POTENTIAL_SHORT_SALE": _sort_key_short_sale,
    "UNCLASSIFIED": _sort_key_unclassified,
}


def build_workbook(results: list[CaseResult]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # default blank sheet

    by_bucket: dict[str, list[CaseResult]] = {b: [] for b in SHEET_ORDER}
    for r in results:
        bucket = r.lead_bucket if r.lead_bucket in by_bucket else "UNCLASSIFIED"
        by_bucket[bucket].append(r)

    summary_col_idx = next(
        i for i, (header, _) in enumerate(COLUMNS, start=1) if header == "Summary of Case"
    )

    for bucket in SHEET_ORDER:
        rows = sorted(by_bucket.get(bucket, []), key=SORT_KEYS[bucket])
        ws = wb.create_sheet(title=bucket)

        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        for row_idx, result in enumerate(rows, start=2):
            for col_idx, (_, getter) in enumerate(COLUMNS, start=1):
                value = getter(result)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == summary_col_idx:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Multi-line bullet text needs a taller row than the default,
            # sized to how many bullets this case actually has -- a fixed
            # height would either clip a HOT case with a long history or
            # waste space on a case with only one or two bullets.
            line_count = result.case_summary.count("\n") + 1 if result.case_summary else 1
            ws.row_dimensions[row_idx].height = max(15, line_count * 15)

        for col_idx in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                60 if col_idx == summary_col_idx else 22
            )
        ws.freeze_panes = "A2"

    return wb


def export_to_xlsx(results: list[CaseResult], output_path: str) -> dict:
    wb = build_workbook(results)
    wb.save(output_path)
    counts = {b: 0 for b in SHEET_ORDER}
    for r in results:
        counts[r.lead_bucket if r.lead_bucket in counts else "UNCLASSIFIED"] += 1
    return counts
