"""The master workbook: Leads / Town Portals / Meta.

The one rule that shapes this whole module: the operator owns seven
columns on the Leads tab (Skip Trace Sent, Phone 1, Phone 2, Email, Call
Attempts, Outcome, My Notes) and the bot must never write to them or
reorder around them. So every write here is *append-only* and
*header-addressed* -- rows are added below the last used row, and each
value is placed by looking its header up in row 1 rather than by column
index. An operator who reorders columns, or inserts one of their own,
does not break the next run and does not get their notes shuffled into
the wrong row.

Existing rows are never rewritten, including when a listing shows up in a
later alert: dedupe happens upstream on the MLS # set returned by
`existing_mls_numbers()`.
"""

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .models import NA, Lead, TownPortal

log = logging.getLogger(__name__)

LEADS_SHEET = "Leads"
PORTALS_SHEET = "Town Portals"
META_SHEET = "Meta"

# Written by the bot. Order here is the order they are created in.
BOT_COLUMNS: list[tuple[str, object]] = [
    ("MLS #", lambda l: l.mls_no),
    ("Status", lambda l: l.alert.status.title() if l.alert.status else NA),
    ("Street Address", lambda l: l.alert.street_address or NA),
    ("Town", lambda l: l.mls.town if l.mls.town != NA else (l.alert.town or NA)),
    ("Zip", lambda l: l.mls.zip_code if l.mls.zip_code != NA else (l.alert.zip_code or NA)),
    ("List Price (Final)", lambda l: l.mls.list_price_final),
    ("Original List Price", lambda l: l.mls.list_price_original),
    ("Days on Market", lambda l: l.mls.days_on_market),
    ("Cumulative DOM", lambda l: l.mls.cumulative_dom),
    ("List Date", lambda l: l.mls.list_date),
    ("Expiration Date", lambda l: l.mls.expiration_date),
    ("Price History", lambda l: l.mls.price_history_text()),
    ("Beds", lambda l: l.mls.beds),
    ("Full Baths", lambda l: l.mls.full_baths),
    ("Half Baths", lambda l: l.mls.half_baths),
    ("Sqft (Above Grade)", lambda l: l.mls.sqft_above_grade),
    ("Lot Size", lambda l: l.mls.lot_size),
    ("Year Built", lambda l: l.mls.year_built),
    ("Property Type", lambda l: l.mls.property_type),
    ("Listing Agent", lambda l: l.mls.listing_agent),
    ("Brokerage", lambda l: l.mls.brokerage),
    ("Public Remarks", lambda l: l.mls.public_remarks),
    ("Owner Name", lambda l: l.owner.owner_name),
    ("Mailing Address", lambda l: l.owner.mailing_address),
    ("Mailing City", lambda l: l.owner.mailing_city),
    ("Mailing State", lambda l: l.owner.mailing_state),
    ("Mailing Zip", lambda l: l.owner.mailing_zip),
    ("Assessed Value", lambda l: l.owner.assessed_value),
    ("Last Sale Date", lambda l: l.owner.last_sale_date),
    ("Last Sale Price", lambda l: l.owner.last_sale_price),
    ("Assessor Living Area", lambda l: l.owner.assessor_living_area),
    ("Assessor Beds", lambda l: l.owner.assessor_beds),
    ("Assessor Baths", lambda l: l.owner.assessor_baths),
    ("Assessor Card URL", lambda l: l.owner.portal_url),
    ("Price Drops", lambda l: l.price_drops),
    ("Total Price Reduction $", lambda l: l.total_reduction_dollars),
    ("Total Price Reduction %", lambda l: l.total_reduction_pct),
    ("Price Per Sqft", lambda l: l.price_per_sqft),
    ("Absentee", lambda l: l.absentee),
    ("Days Since Expired", lambda l: l.days_since_expired),
    ("Lead Score", lambda l: l.lead_score),
    ("MLS Pull Status", lambda l: l.mls.pull_status),
    ("Notes", lambda l: "; ".join(filter(None, [l.owner.notes_text(), l.notes]))),
    ("First Seen", lambda l: datetime.now(timezone.utc).date().isoformat()),
]

# Owned by the operator. Created empty, then never touched again.
MANUAL_COLUMNS = [
    "Skip Trace Sent", "Phone 1", "Phone 2", "Email",
    "Call Attempts", "Outcome", "My Notes",
]

PORTAL_COLUMNS = ["Town", "Portal URL", "Platform", "Search Notes"]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def create_workbook(path: str | Path) -> None:
    """First-run setup: three tabs, headers only."""
    wb = Workbook()
    wb.remove(wb.active)

    leads = wb.create_sheet(LEADS_SHEET)
    headers = [name for name, _ in BOT_COLUMNS] + MANUAL_COLUMNS
    for idx, header in enumerate(headers, start=1):
        cell = leads.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        leads.column_dimensions[get_column_letter(idx)].width = (
            60 if header in ("Public Remarks", "Price History", "Notes") else 20
        )
    leads.freeze_panes = "A2"

    portals = wb.create_sheet(PORTALS_SHEET)
    for idx, header in enumerate(PORTAL_COLUMNS, start=1):
        portals.cell(row=1, column=idx, value=header).font = Font(bold=True)
        portals.column_dimensions[get_column_letter(idx)].width = 46 if idx == 4 else 26
    portals.freeze_panes = "A2"

    meta = wb.create_sheet(META_SHEET)
    meta.cell(row=1, column=1, value="key").font = Font(bold=True)
    meta.cell(row=1, column=2, value="value").font = Font(bold=True)
    meta.cell(row=2, column=1, value="last_run")
    meta.cell(row=2, column=2, value="")
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 34

    wb.save(path)
    log.info("created master workbook at %s", path)


def ensure_workbook(path: str | Path) -> bool:
    """Create the workbook if absent. Returns True if it was created."""
    if Path(path).exists():
        return False
    create_workbook(path)
    return True


def _header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


def existing_mls_numbers(path: str | Path) -> set[str]:
    """Every MLS # already on the Leads tab -- the dedupe set for Step 1."""
    if not Path(path).exists():
        return set()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if LEADS_SHEET not in wb.sheetnames:
            return set()
        sheet = wb[LEADS_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return set()
        try:
            mls_idx = [str(h).strip() if h else "" for h in header].index("MLS #")
        except ValueError:
            return set()
        return {
            str(row[mls_idx]).strip()
            for row in rows
            if row and mls_idx < len(row) and row[mls_idx] not in (None, "")
        }
    finally:
        wb.close()


def read_last_run(path: str | Path) -> str:
    if not Path(path).exists():
        return ""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if META_SHEET not in wb.sheetnames:
            return ""
        for row in wb[META_SHEET].iter_rows(values_only=True):
            if row and str(row[0]).strip() == "last_run":
                return str(row[1]).strip() if len(row) > 1 and row[1] else ""
        return ""
    finally:
        wb.close()


def read_portals(path: str | Path) -> dict[str, TownPortal]:
    if not Path(path).exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if PORTALS_SHEET not in wb.sheetnames:
            return {}
        rows = wb[PORTALS_SHEET].iter_rows(values_only=True)
        next(rows, None)
        registry: dict[str, TownPortal] = {}
        for row in rows:
            if not row or not row[0]:
                continue
            registry[str(row[0]).strip()] = TownPortal(
                town=str(row[0]).strip(),
                portal_url=str(row[1] or "").strip(),
                platform=str(row[2] or "Unknown").strip(),
                search_notes=str(row[3] or "").strip(),
            )
        return registry
    finally:
        wb.close()


def append_leads(path: str | Path, leads: list[Lead]) -> int:
    """Append rows to the Leads tab, header-addressed. Returns rows added.

    Any bot column missing from an operator-modified sheet is appended as
    a new header rather than silently dropped, and manual columns are
    never written -- the row is simply left blank under them.
    """
    if not leads:
        return 0
    wb = load_workbook(path)
    try:
        sheet = wb[LEADS_SHEET]
        headers = _header_map(sheet)

        for name, _ in BOT_COLUMNS:
            if name not in headers:
                new_col = (max(headers.values()) if headers else 0) + 1
                sheet.cell(row=1, column=new_col, value=name).font = Font(bold=True)
                headers[name] = new_col

        row_idx = sheet.max_row + 1
        for lead in leads:
            for name, getter in BOT_COLUMNS:
                cell = sheet.cell(row=row_idx, column=headers[name], value=getter(lead))
                if name in ("Public Remarks", "Price History", "Notes"):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

        wb.save(path)
        return len(leads)
    finally:
        wb.close()


def upsert_portals(path: str | Path, portals: list[TownPortal]) -> int:
    """Add or update Town Portals rows. Returns rows written."""
    if not portals:
        return 0
    wb = load_workbook(path)
    try:
        sheet = wb[PORTALS_SHEET]
        existing: dict[str, int] = {}
        for row in range(2, sheet.max_row + 1):
            town = sheet.cell(row=row, column=1).value
            if town:
                existing[str(town).strip()] = row

        written = 0
        for portal in portals:
            target = existing.get(portal.town, sheet.max_row + 1)
            sheet.cell(row=target, column=1, value=portal.town)
            sheet.cell(row=target, column=2, value=portal.portal_url)
            sheet.cell(row=target, column=3, value=portal.platform)
            sheet.cell(row=target, column=4, value=portal.search_notes)
            existing[portal.town] = target
            written += 1

        wb.save(path)
        return written
    finally:
        wb.close()


def write_last_run(path: str | Path, timestamp: str | None = None) -> None:
    wb = load_workbook(path)
    try:
        sheet = wb[META_SHEET]
        for row in range(1, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=1).value).strip() == "last_run":
                sheet.cell(row=row, column=2, value=timestamp or _now_iso())
                wb.save(path)
                return
        target = sheet.max_row + 1
        sheet.cell(row=target, column=1, value="last_run")
        sheet.cell(row=target, column=2, value=timestamp or _now_iso())
        wb.save(path)
    finally:
        wb.close()
