"""Read a connectMLS "Export Listings" file. Replaces Steps 1 and 2.

This is the best input the bot has, and it should be the default one.
Exporting the saved search from connectMLS gives typed, structured rows
straight from the MLS -- no email parsing, no OAuth, no API
reverse-engineering, no HTML scraping, and nothing that breaks when
connectMLS changes its markup.

CONFIRMED against a real export, 2026-08-23:
`connectMLS_Skip_trace_20260823172408.XLS`, 103 rows, legacy .xls
(OLE compound), columns:

    Photo, Proximity, MLS#, Status, Status Timestamp, Prop Type,
    List/Closed Price, Address, City, Acres, SQFT Total, SQFTAbvGr,
    Style, Rooms, Beds, Baths, Zip Code, Garage/Park, Built, DOM,
    LstAgt/Team, List Office W/ B Card

FOUR THINGS THE EXPORT DOES THAT WILL BITE YOU IF UNHANDLED, all
observed in that file and all handled below:

1. ZIP CODES LOSE THEIR LEADING ZERO. Excel stores them as numbers, so
   06118 arrives as the float 6118.0. Every Connecticut ZIP starts with
   0, so every single row is affected -- and a 4-digit ZIP breaks the
   assessor lookup and ships malformed rows to the skip trace vendor.
   Zero-padded back to 5 digits here.

2. MLS NUMBERS ARE FLOATS TOO. 24186595.0, not "24186595". Left as-is
   they never match the dedupe set, so every run re-processes every row.

3. PRICES CARRY A PREFIX: "LP: $339,000".

4. RENTALS ARE MIXED IN WITH SALES. The real export had a listing at
   "LP: $2,375" -- a monthly rent, not a sale price. Scored naively that
   is a $2.46/sqft property. Flagged as `likely_rental` and routed to
   review rather than silently priced as a sale.

WHAT THE EXPORT DOES NOT CARRY: Original List Price, Cumulative DOM,
List Date, price history, and public remarks. Original List Price is the
notable loss -- without it there is no price-reduction figure at all, so
`Total Price Reduction` stays N/A and price-drop scoring falls back to
unknown. Adding "Original List Price" to the export template in
connectMLS would close that gap; see the module docstring note in
scoring.py for why unknown is not the same as zero.

`Status Timestamp` is the date the listing changed to its current
status. For an EXPD row that is when it expired, which is what
"Days Since Expired" needs -- but it is the status-change date, not the
MLS's own "Expiration Date" field, and the two can differ by a day.
"""

import csv
import logging
import re
from pathlib import Path

from .models import NA, AlertListing, MlsDetail

log = logging.getLogger(__name__)

# Below this, a "list price" is a monthly rent rather than a sale price.
# Nothing habitable in CT sells for under $20k; the observed rental was
# $2,375. Rows under the threshold are flagged, never silently dropped.
RENTAL_PRICE_CEILING = 20000

# Export column -> what it means here. Matched case-insensitively and
# whitespace-collapsed, so a template with slightly different spacing
# still lines up.
COLUMN_ALIASES = {
    "mls#": "mls_no",
    "mls #": "mls_no",
    "status": "status",
    "status timestamp": "status_timestamp",
    "prop type": "prop_type",
    "style": "style",
    "list/closed price": "list_price",
    "list price": "list_price",
    "original list price": "original_list_price",
    "address": "address",
    "city": "city",
    "zip code": "zip_code",
    "zip": "zip_code",
    "acres": "acres",
    "sqft total": "sqft_total",
    "sqftabvgr": "sqft_above_grade",
    "sqft abv gr": "sqft_above_grade",
    "beds": "beds",
    "baths": "baths",
    "rooms": "rooms",
    "built": "year_built",
    "year built": "year_built",
    "dom": "dom",
    "cdom": "cdom",
    "lstagt/team": "list_agent",
    "list agent": "list_agent",
    "list office w/ b card": "list_office",
    "list office": "list_office",
    "expiration date": "expiration_date",
    "public remarks": "public_remarks",
    "remarks": "public_remarks",
}

_STATUS_MAP = {
    "expd": "expired", "exp": "expired", "expired": "expired",
    "with": "withdrawn", "wdrn": "withdrawn", "withdrawn": "withdrawn",
    "canc": "canceled", "cncl": "canceled", "canceled": "canceled",
    "cancelled": "canceled",
    "tom": "temporarily off market", "temp off market": "temporarily off market",
}


def _norm_header(raw) -> str:
    return re.sub(r"\s+", " ", str(raw or "").strip()).lower()


def _clean(value) -> str:
    """Cell value -> trimmed string. Floats that are whole become ints."""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_zip(value) -> str:
    """6118.0 -> '06118'. See point 1 in the module docstring."""
    text = _clean(value)
    if not text:
        return ""
    digits = re.sub(r"[^0-9]", "", text.split("-")[0])
    if not digits:
        return ""
    return digits.zfill(5)[:5]


def parse_price(value) -> str:
    """'LP: $339,000' -> '339000'. Unparseable -> NA."""
    text = _clean(value)
    digits = re.sub(r"[^0-9.]", "", text)
    if not digits:
        return NA
    try:
        return str(int(float(digits)))
    except ValueError:
        return NA


def split_baths(value) -> tuple[str, str]:
    """'2/1' -> ('2', '1'). A bare '2' -> ('2', NA), not ('2', '0')."""
    text = _clean(value)
    if not text:
        return NA, NA
    if "/" in text:
        full, _, half = text.partition("/")
        return (_clean(full) or NA), (_clean(half) or NA)
    return text, NA


def normalize_status(value) -> str:
    text = _clean(value).lower()
    return _STATUS_MAP.get(text, text)


def strip_unit(address: str) -> str:
    """'16 Greenbriar Drive, Unit# E' -> '16 Greenbriar Drive'.

    The assessor search wants the street address; a unit suffix makes it
    miss. When several parcels then match, assessor.py disambiguates on
    year built and living area, and refuses rather than guessing.

    Real exports carry three shapes, all observed in the same file:
    "Unit# E", "Unit# LOT 12" (multi-word), and a bare "Unit#" with
    nothing after it. The unit value is therefore optional and may run to
    several tokens -- requiring exactly one token silently left the
    marker on two thirds of the condo rows.
    """
    cleaned = re.sub(r"\s+", " ", (address or "").strip())
    cleaned = re.split(
        r",?\s*(?:unit\s*#?|apt\.?|ste\.?|suite|#)\s*.*$", cleaned, flags=re.IGNORECASE
    )[0]
    return cleaned.strip(" ,")


_UNIT_RE = re.compile(
    r",?\s*(?:unit\s*#?|apt\.?|ste\.?|suite|#)\s*(.*)$", re.IGNORECASE
)


def extract_unit(address: str) -> str:
    """'16 Greenbriar Drive, Unit# E' -> 'E'. Bare 'Unit#' -> ''."""
    match = _UNIT_RE.search(re.sub(r"\s+", " ", (address or "").strip()))
    return match.group(1).strip(" ,#") if match else ""


def _property_type(row: dict) -> str:
    parts = [row.get("prop_type", ""), row.get("style", "")]
    joined = " - ".join(p for p in parts if p)
    return joined or NA


def _lot_size(row: dict) -> str:
    acres = row.get("acres", "")
    return f"{acres} acres" if acres else NA


def row_to_records(row: dict) -> tuple[AlertListing, MlsDetail] | None:
    """One export row -> (alert, detail). None if it has no MLS number."""
    mls_no = _clean(row.get("mls_no"))
    if not mls_no:
        return None

    zip_code = normalize_zip(row.get("zip_code"))
    full_address = _clean(row.get("address"))
    status = normalize_status(row.get("status"))

    alert = AlertListing(
        mls_no=mls_no,
        street_address=strip_unit(full_address),
        unit=extract_unit(full_address),
        town=_clean(row.get("city")),
        zip_code=zip_code,
        status=status,
        source_email="connectmls-export",
    )

    full_baths, half_baths = split_baths(row.get("baths"))
    price = parse_price(row.get("list_price"))

    detail = MlsDetail(mls_no=mls_no)
    detail.list_price_final = price
    detail.list_price_original = parse_price(row.get("original_list_price"))
    detail.days_on_market = _clean(row.get("dom")) or NA
    detail.cumulative_dom = _clean(row.get("cdom")) or NA
    # See the docstring: this is the status-change date, which for an
    # expired row is when it expired.
    detail.expiration_date = (
        _clean(row.get("expiration_date")) or _clean(row.get("status_timestamp")) or NA
    )
    detail.beds = _clean(row.get("beds")) or NA
    detail.full_baths = full_baths
    detail.half_baths = half_baths
    detail.sqft_above_grade = (
        _clean(row.get("sqft_above_grade")) or _clean(row.get("sqft_total")) or NA
    )
    detail.lot_size = _lot_size(row)
    detail.year_built = _clean(row.get("year_built")) or NA
    detail.property_type = _property_type(row)
    detail.town = alert.town or NA
    detail.zip_code = zip_code or NA
    detail.listing_agent = _clean(row.get("list_agent")) or NA
    detail.brokerage = _clean(row.get("list_office")) or NA
    remarks = _clean(row.get("public_remarks"))
    detail.public_remarks = remarks[:200] if remarks else NA

    numeric_price = None
    if price != NA:
        try:
            numeric_price = float(price)
        except ValueError:
            numeric_price = None
    if numeric_price is not None and numeric_price < RENTAL_PRICE_CEILING:
        detail.likely_rental = True

    return alert, detail


def _rows_from_xls(path: Path) -> list[dict]:
    import xlrd  # only needed for legacy .xls

    sheet = xlrd.open_workbook(str(path)).sheet_by_index(0)
    headers = [_norm_header(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
    rows = []
    for r in range(1, sheet.nrows):
        row = {}
        for c, header in enumerate(headers):
            key = COLUMN_ALIASES.get(header)
            if key:
                row[key] = sheet.cell_value(r, c)
        rows.append(row)
    return rows


def _rows_from_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    sheet = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    iterator = sheet.iter_rows(values_only=True)
    headers = [_norm_header(h) for h in (next(iterator, []) or [])]
    rows = []
    for values in iterator:
        row = {}
        for header, value in zip(headers, values):
            key = COLUMN_ALIASES.get(header)
            if key:
                row[key] = value
        rows.append(row)
    return rows


def _rows_from_csv(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for raw in reader:
            row = {}
            for header, value in raw.items():
                key = COLUMN_ALIASES.get(_norm_header(header))
                if key:
                    row[key] = value
            rows.append(row)
    return rows


def load_export(path: str | Path) -> list[tuple[AlertListing, MlsDetail]]:
    """Read a connectMLS export (.xls, .xlsx or .csv) into records."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        raw_rows = _rows_from_xls(path)
    elif suffix in (".xlsx", ".xlsm"):
        raw_rows = _rows_from_xlsx(path)
    elif suffix in (".csv", ".txt"):
        raw_rows = _rows_from_csv(path)
    else:
        raise ValueError(f"Unsupported export format {suffix!r} (expected .xls/.xlsx/.csv)")

    records = []
    for raw in raw_rows:
        record = row_to_records(raw)
        if record:
            records.append(record)
    log.info("loaded %d listings from %s", len(records), path.name)
    return records


def expired_only(records: list[tuple[AlertListing, MlsDetail]]):
    """Keep the statuses this bot is for; drop Active/Pending/Sold."""
    from .models import EXPIRED_STATUSES

    return [(a, d) for a, d in records if a.status in EXPIRED_STATUSES]
