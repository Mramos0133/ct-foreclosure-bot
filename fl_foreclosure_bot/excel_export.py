"""Builds a ranked, multi-sheet Excel workbook from scraped listings.

Sheets, in tab order (per explicit request -- the active/upcoming
auctions are the actionable leads, so they come first and aren't mixed
in with the noise of already-decided ones):

  ACTIVE    -- anything not yet sold or canceled: scheduled/waiting/
               running auctions. Sorted soonest auction date first, then
               biggest equity cushion first within a date.
  SOLD      -- status "Sold". Useful as comp/outcome data.
  CANCELED  -- status contains "cancel" (e.g. "Canceled per Order",
               "Canceled per Bankruptcy"). Kept for reference; a
               canceled sale can also signal a workout/short-sale
               conversation happening -- but per explicit request these
               are out of the main view.

Bucketing keys off the auction_status text since that's all this data
source exposes (no docket detail here, unlike the CT bot). An empty or
unrecognized status lands in ACTIVE rather than being hidden away --
better to over-show than silently bury a live lead.

equity_estimate (assessed value minus final judgment) drives the
within-date ranking; rows missing either figure sort last, not dropped.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import AuctionListing

COLUMNS = [
    ("County", lambda r: r.county),
    ("Auction Date", lambda r: r.auction_date),
    ("Section", lambda r: r.section),
    ("Auction Status", lambda r: r.auction_status),
    ("Auction Start Time", lambda r: r.auction_start_time or ""),
    ("Sold Date/Time", lambda r: r.sold_date_time or ""),
    ("Sold Amount", lambda r: r.sold_amount if r.sold_amount is not None else ""),
    ("Sold To", lambda r: r.sold_to or ""),
    ("Auction Type", lambda r: r.auction_type),
    ("Case No.", lambda r: r.case_no),
    ("Case URL", lambda r: r.case_no_url or ""),
    ("Final Judgment Amount", lambda r: r.final_judgment_amount if r.final_judgment_amount is not None else ""),
    ("Parcel ID", lambda r: r.parcel_id or ""),
    ("Parcel URL", lambda r: r.parcel_id_url or ""),
    ("Property Address", lambda r: r.property_address),
    ("Assessed Value", lambda r: r.assessed_value if r.assessed_value is not None else ""),
    ("Equity Estimate", lambda r: r.equity_estimate if r.equity_estimate is not None else ""),
    ("Plaintiff Max Bid", lambda r: r.plaintiff_max_bid),
    ("Source URL", lambda r: r.source_url),
]


SHEET_ORDER = ["ACTIVE", "SOLD", "CANCELED"]


def bucket_for(r: AuctionListing) -> str:
    status = r.auction_status.strip().lower()
    if status == "sold":
        return "SOLD"
    if "cancel" in status:
        return "CANCELED"
    return "ACTIVE"


def _sort_key_active(r: AuctionListing):
    equity = r.equity_estimate
    return (r.auction_date, equity is None, -(equity or 0))


def _sort_key_decided(r: AuctionListing):
    equity = r.equity_estimate
    return (equity is None, -(equity or 0))


SORT_KEYS = {
    "ACTIVE": _sort_key_active,
    "SOLD": _sort_key_decided,
    "CANCELED": _sort_key_decided,
}


def build_workbook(results: list[AuctionListing]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # default blank sheet

    by_bucket: dict[str, list[AuctionListing]] = {b: [] for b in SHEET_ORDER}
    for r in results:
        by_bucket[bucket_for(r)].append(r)

    for bucket in SHEET_ORDER:
        ws = wb.create_sheet(title=bucket)
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        rows = sorted(by_bucket[bucket], key=SORT_KEYS[bucket])
        for row_idx, result in enumerate(rows, start=2):
            for col_idx, (_, getter) in enumerate(COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=getter(result))

        for col_idx in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.freeze_panes = "A2"
    return wb


def export_to_xlsx(results: list[AuctionListing], output_path: str) -> dict[str, int]:
    wb = build_workbook(results)
    wb.save(output_path)
    counts = {b: 0 for b in SHEET_ORDER}
    for r in results:
        counts[bucket_for(r)] += 1
    return counts
